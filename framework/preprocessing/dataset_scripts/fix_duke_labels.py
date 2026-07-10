from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections.abc import Sequence
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[4]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))


DATASET_ID = "duke"
DEFAULT_DUKE_ROOT = Path("F:/open_data/breast/Duke")
DEFAULT_DATASET_ROOT = Path(os.environ.get("BREAST_MRI_DATASET_ROOT", "/home/ubuntu/liuyiyao1/multimodal_dataset"))
DEFAULT_PROCESSED_ROOT = DEFAULT_DATASET_ROOT / "full_v4_duke"
DEFAULT_CLINICAL_FILE = "Clinical_and_Other_Features.xlsx"

PATIENT_ID = "Patient ID"
PCR_COLUMN = "Pathologic Response to Neoadjuvant Therapy"

MISSING_TOKENS = {"", "NA", "N/A", "NC", "NP", "NONE", "NULL", "NAN"}

RECEPTOR_MAP = {0: "negative", 1: "positive", 2: "equivocal"}
YES_NO_MAP = {0: "no", 1: "yes", 2: "no"}
MENOPAUSE_MAP = {0: "pre", 1: "post", 2: "not_applicable"}
RACE_ETHNICITY_MAP = {
    0: "not_applicable",
    1: "white",
    2: "black",
    3: "asian",
    4: "native",
    5: "hispanic",
    6: "multi",
    7: "hawaiian_pacific_islander",
    8: "american_indian",
}
MOLECULAR_SUBTYPE_MAP = {
    0: "Luminal-like",
    1: "HR+/HER2+",
    2: "HER2-enriched",
    3: "Triple-negative",
}
HISTOLOGY_MAP = {
    0: "DCIS",
    1: "ductal",
    2: "lobular",
    3: "metaplastic",
    4: "LCIS",
    5: "tubular",
    6: "mixed",
    7: "micropapillary",
    8: "colloid",
    9: "mucinous",
    10: "medullary",
}
SURGERY_TYPE_MAP = {0: "breast_conserving_surgery", 1: "mastectomy"}
CLINICAL_RESPONSE_MAP = {
    1: "complete_response_on_imaging",
    2: "not_complete_response",
    3: "imaging_unavailable",
}
PATHOLOGIC_RESPONSE_MAP = {
    1: "complete_response",
    2: "not_complete_response",
    3: "dcis_only_remaining",
    4: "lcis_only_remaining",
    5: "assessment_unavailable",
}
NEAR_COMPLETE_MAP = {
    0: "not_complete_or_near_complete",
    1: "complete",
    2: "near_complete",
    3: "pathologic_assessment_unavailable",
}

T_STAGE_MAP = {-1: "TX", 0: "T0", 1: "T1", 2: "T2", 3: "T3", 4: "T4", 5: "Tis"}
N_STAGE_MAP = {-1: "NX", 0: "N0", 1: "N1", 2: "N2", 3: "N3"}
M_STAGE_MAP = {-1: "MX", 0: "M0", 1: "M1"}

DUKE_LABEL_SUMMARY_FIELDS = [
    "dataset_id",
    "subject_id",
    "timepoint",
    "study_uid",
    "labels_json",
    "pCR",
    "pcr_status",
    "HER2",
    "ER",
    "PR",
    "molecular_subtype",
    "laterality",
    "age",
    "BMI",
    "source_label_count",
    "split",
]


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Repair Duke labels in an already-built multimodal dataset. "
            "The script rebuilds labels from Clinical_and_Other_Features.xlsx "
            "and overwrites Duke labels.json files only when executed."
        )
    )
    parser.add_argument("--duke-root", type=Path, default=DEFAULT_DUKE_ROOT)
    parser.add_argument("--clinical-xlsx", type=Path, default=None)
    parser.add_argument("--processed-root", type=Path, default=DEFAULT_PROCESSED_ROOT)
    parser.add_argument("--dataset-id", default=DATASET_ID)
    parser.add_argument("--dry-run", default=False, help="Analyze and print counts without writing files.")
    parser.add_argument("--backup", action="store_true", help="Create labels.json.duke_label_fix.bak before overwriting.")
    args = parser.parse_args(argv)

    clinical_xlsx = args.clinical_xlsx or args.duke_root / DEFAULT_CLINICAL_FILE
    labels_by_subject = load_duke_clinical_labels(clinical_xlsx, dataset_id=args.dataset_id)
    summary = repair_processed_duke_labels(
        processed_root=args.processed_root,
        labels_by_subject=labels_by_subject,
        clinical_xlsx=clinical_xlsx,
        dataset_id=args.dataset_id,
        dry_run=args.dry_run,
        backup=args.backup,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def load_duke_clinical_labels(clinical_xlsx: Path, dataset_id: str = DATASET_ID) -> dict[str, dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("openpyxl is required to read Duke Clinical_and_Other_Features.xlsx.") from exc

    if not clinical_xlsx.exists():
        raise FileNotFoundError(f"Missing Duke clinical label file: {clinical_xlsx}")

    workbook = load_workbook(clinical_xlsx, read_only=True, data_only=True)
    try:
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.worksheets[0]
        header_values = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
        headers = _unique_headers(header_values)
        labels_by_subject: dict[str, dict[str, Any]] = {}
        for row_values in worksheet.iter_rows(min_row=4, values_only=True):
            row = _row_dict(headers, row_values)
            subject_id = normalize_duke_subject_id(row.get(PATIENT_ID, ""))
            if not subject_id:
                continue
            labels_by_subject[subject_id] = build_duke_label_payload(
                subject_id=subject_id,
                row=row,
                clinical_xlsx=clinical_xlsx,
                dataset_id=dataset_id,
            )
        return labels_by_subject
    finally:
        workbook.close()


def build_duke_label_payload(
    subject_id: str,
    row: dict[str, Any],
    clinical_xlsx: Path,
    dataset_id: str = DATASET_ID,
) -> dict[str, Any]:
    pcr_value = normalize_duke_pcr(row.get(PCR_COLUMN))
    pcr_status = str(pcr_value) if pcr_value in {0, 1} else "unknown"
    er = decode_receptor(row.get("ER"))
    pr = decode_receptor(row.get("PR"))
    her2 = decode_receptor(row.get("HER2"))
    molecular_subtype = decode_map(row.get("Mol Subtype"), MOLECULAR_SUBTYPE_MAP)
    baseline_stage = stage_summary(
        row.get("Staging(Tumor Size)# [T]"),
        row.get("Staging(Nodes)#(Nx replaced by -1)[N]"),
        row.get("Staging(Metastasis)#(Mx -replaced by -1)[M]"),
    )
    post_nact_stage = stage_summary(
        row.get("Pathologic response to Neoadjuvant therapy: Pathologic stage (T) following neoadjuvant therapy "),
        row.get("Pathologic response to Neoadjuvant therapy:  Pathologic stage (N) following neoadjuvant therapy"),
        row.get("Pathologic response to Neoadjuvant therapy:  Pathologic stage (M) following neoadjuvant therapy "),
    )
    survival = build_survival_payload(row)
    treatment = build_treatment_payload(row)
    response = {
        "clinical_response_imaging": decode_map(
            row.get("Clinical Response, Evaluated Through Imaging "),
            CLINICAL_RESPONSE_MAP,
        ),
        "pathologic_response": decode_map(row.get(PCR_COLUMN), PATHOLOGIC_RESPONSE_MAP),
        "near_complete_strict": decode_map(
            row.get("Overall Near-complete Response:  Stricter Definition"),
            NEAR_COMPLETE_MAP,
        ),
        "near_complete_loose": decode_map(
            row.get("Overall Near-complete Response:  Looser Definition"),
            NEAR_COMPLETE_MAP,
        ),
        "near_complete_graded": clean_value(row.get("Near-complete Response (Graded Measure)")),
    }

    payload = {
        "schema_version": "unified_labels_v1",
        "dataset_id": dataset_id,
        "subject_id": subject_id,
        "patient_id": subject_id,
        "timepoint": "unknown",
        "study_uid": "unknown",
        "pCR": pcr_value,
        "pcr_status": pcr_status,
        "HER2": her2,
        "ER": er,
        "PR": pr,
        "molecular_subtype": molecular_subtype,
        "age": age_years_from_birth_days(row.get("Date of Birth (Days)")),
        "BMI": None,
        "sex": "female",
        "menopause": decode_map(row.get("Menopause (at diagnosis)"), MENOPAUSE_MAP),
        "race_ethnicity": decode_map(row.get("Race and Ethnicity"), RACE_ETHNICITY_MAP),
        "laterality": normalize_laterality(row.get("Tumor Location")),
        "bilateral": decode_yes_no(row.get("Bilateral Information")),
        "clinical_stage": baseline_stage["summary"],
        "pathologic_stage": post_nact_stage["summary"],
        "tumor_size_cm": first_non_missing(row.get("Tumor Size (cm)"), row.get("Tumor Size (cm).1")),
        "oncotype_score": numeric_or_none(row.get("Oncotype score")),
        "metastatic_at_presentation": decode_yes_no(row.get("Metastatic at Presentation (Outside of Lymph Nodes)")),
        "pathology": {
            "grade": numeric_or_none(row.get("Tumor Grade")),
            "nottingham_grade": numeric_or_none(row.get("Nottingham grade")),
            "histologic_type": decode_map(row.get("Histologic type"), HISTOLOGY_MAP),
            "baseline_t": baseline_stage["t"],
            "baseline_n": baseline_stage["n"],
            "baseline_m": baseline_stage["m"],
            "post_neoadjuvant_t": post_nact_stage["t"],
            "post_neoadjuvant_n": post_nact_stage["n"],
            "post_neoadjuvant_m": post_nact_stage["m"],
        },
        "treatment": treatment,
        "response": response,
        "survival": survival,
        "standardized_sources": {
            "pCR": [PCR_COLUMN],
            "HER2": ["HER2"],
            "ER": ["ER"],
            "PR": ["PR"],
            "molecular_subtype": ["Mol Subtype"],
            "survival": [
                "Recurrence event(s)",
                "Days to local recurrence (from the date of diagnosis) ",
                "Days to distant recurrence(from the date of diagnosis) ",
                "Days to death (from the date of diagnosis) ",
                "Age at last contact in EMR f/u(days)(from the date of diagnosis) ,last time patient known to be alive, unless age of death is reported(in such case the age of death",
            ],
        },
        "source_label_count": 1,
        "source_files": [str(clinical_xlsx)],
        "duke_core_labels": {},
        "raw_label_records": [
            {
                "source_file": str(clinical_xlsx),
                "key_column": PATIENT_ID,
                "labels": json_safe(row),
                "notes": "Duke clinical labels rebuilt with dataset-specific pCR rule.",
            }
        ],
    }
    payload["duke_core_labels"] = compact_core_labels(payload)
    return json_safe(payload)


def repair_processed_duke_labels(
    processed_root: Path,
    labels_by_subject: dict[str, dict[str, Any]],
    clinical_xlsx: Path,
    dataset_id: str = DATASET_ID,
    dry_run: bool = False,
    backup: bool = False,
) -> dict[str, Any]:
    processed_root = processed_root.expanduser().resolve(strict=False)
    if not processed_root.exists():
        raise FileNotFoundError(f"Missing processed Duke dataset root: {processed_root}")

    label_paths = sorted(path for path in processed_root.rglob("labels.json") if is_duke_label_path(path, dataset_id))
    updated_labels = 0
    missing_subjects: list[str] = []
    labels_summary_rows: list[dict[str, Any]] = []

    for labels_path in label_paths:
        existing = read_json_object(labels_path)
        subject_id = normalize_duke_subject_id(
            existing.get("subject_id")
            or existing.get("patient_id")
            or infer_subject_from_label_path(labels_path, dataset_id)
        )
        if not subject_id or subject_id not in labels_by_subject:
            if subject_id:
                missing_subjects.append(subject_id)
            continue
        repaired = merge_existing_label(
            existing=existing,
            repaired=labels_by_subject[subject_id],
            clinical_xlsx=clinical_xlsx,
        )
        if not dry_run:
            if backup:
                backup_path = labels_path.with_name(labels_path.name + ".duke_label_fix.bak")
                if not backup_path.exists():
                    backup_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
            write_json_atomic(labels_path, repaired)
        updated_labels += 1
        labels_summary_rows.append(label_summary_row(labels_path, repaired))

    updated_csvs = update_processed_csvs(
        processed_root=processed_root,
        labels_by_subject=labels_by_subject,
        labels_summary_rows=labels_summary_rows,
        dataset_id=dataset_id,
        dry_run=dry_run,
    )
    return {
        "clinical_xlsx": str(clinical_xlsx),
        "processed_root": str(processed_root),
        "subjects_in_clinical_file": len(labels_by_subject),
        "duke_label_json_files_found": len(label_paths),
        "duke_label_json_files_updated": updated_labels,
        "missing_subjects_in_clinical_file": sorted(set(missing_subjects))[:50],
        "missing_subject_count": len(set(missing_subjects)),
        "csv_files_updated": updated_csvs,
        "dry_run": dry_run,
    }


def update_processed_csvs(
    processed_root: Path,
    labels_by_subject: dict[str, dict[str, Any]],
    labels_summary_rows: list[dict[str, Any]],
    dataset_id: str,
    dry_run: bool,
) -> list[str]:
    updated: list[str] = []
    for name in ("inference_manifest.csv", "training_manifest.csv", "labels_summary.csv"):
        for csv_path in sorted(processed_root.rglob(name)):
            if name == "labels_summary.csv" and labels_summary_rows:
                changed = update_labels_summary_csv(csv_path, labels_summary_rows, dataset_id, dry_run=dry_run)
                if changed:
                    updated.append(str(csv_path))
                continue
            changed = update_manifest_like_csv(csv_path, labels_by_subject, dataset_id, dry_run=dry_run)
            if changed:
                updated.append(str(csv_path))

    for inference_path in sorted(processed_root.rglob("inference_manifest.csv")):
        training_path = inference_path.with_name("training_manifest.csv")
        rows, fieldnames = read_csv_rows(inference_path)
        if not rows:
            continue
        training_rows = [row for row in rows if row.get("pcr_status") in {"0", "1"}]
        if not dry_run:
            write_csv_atomic(training_path, training_rows, fieldnames)
        updated.append(str(training_path))
    return sorted(set(updated))


def update_labels_summary_csv(
    csv_path: Path,
    repaired_rows: list[dict[str, Any]],
    dataset_id: str,
    dry_run: bool,
) -> bool:
    existing_rows, existing_fields = read_csv_rows(csv_path)
    fields = list(existing_fields or DUKE_LABEL_SUMMARY_FIELDS)
    for field in DUKE_LABEL_SUMMARY_FIELDS:
        if field not in fields:
            fields.append(field)
    repaired_by_key = {
        (
            row.get("dataset_id", dataset_id),
            row.get("subject_id", ""),
            row.get("timepoint", "unknown"),
            row.get("study_uid", "unknown"),
        ): row
        for row in repaired_rows
    }

    changed = False
    output_rows: list[dict[str, Any]] = []
    seen_keys: set[tuple[Any, ...]] = set()
    for row in existing_rows:
        key = (
            row.get("dataset_id", ""),
            row.get("subject_id", ""),
            row.get("timepoint", "unknown"),
            row.get("study_uid", "unknown"),
        )
        if row.get("dataset_id") == dataset_id and key in repaired_by_key:
            output_rows.append(repaired_by_key[key])
            seen_keys.add(key)
            changed = True
        else:
            output_rows.append(row)

    for key, row in repaired_by_key.items():
        if key in seen_keys:
            continue
        output_rows.append(row)
        changed = True

    if changed and not dry_run:
        write_csv_atomic(csv_path, output_rows, fields)
    return changed


def update_manifest_like_csv(
    csv_path: Path,
    labels_by_subject: dict[str, dict[str, Any]],
    dataset_id: str,
    dry_run: bool,
) -> bool:
    rows, fieldnames = read_csv_rows(csv_path)
    if not rows:
        return False
    changed = False
    for row in rows:
        if row.get("dataset_id") not in {"", dataset_id}:
            continue
        subject_id = normalize_duke_subject_id(row.get("subject_id") or row.get("patient_uid") or row.get("patient_id"))
        labels = labels_by_subject.get(subject_id)
        if not labels:
            continue
        update_row_with_public_labels(row, labels)
        changed = True
    if changed and not dry_run:
        write_csv_atomic(csv_path, rows, fieldnames)
    return changed


def merge_existing_label(
    existing: dict[str, Any],
    repaired: dict[str, Any],
    clinical_xlsx: Path,
) -> dict[str, Any]:
    merged = dict(existing)
    preserved_keys = {
        "timepoint": existing.get("timepoint", repaired.get("timepoint", "unknown")),
        "study_uid": existing.get("study_uid", repaired.get("study_uid", "unknown")),
        "sample_dir": existing.get("sample_dir"),
        "split": existing.get("split"),
        "imaging_features": existing.get("imaging_features", {}),
    }
    merged.update(repaired)
    for key, value in preserved_keys.items():
        if value not in (None, ""):
            merged[key] = value

    old_records = existing.get("raw_label_records", [])
    if isinstance(old_records, list):
        clinical_name = clinical_xlsx.name.lower()
        non_clinical_records = [
            record
            for record in old_records
            if clinical_name not in str(record.get("source_file", "")).lower()
        ]
        merged["raw_label_records"] = [*non_clinical_records, *repaired.get("raw_label_records", [])]
        merged["source_label_count"] = len(merged["raw_label_records"])

    source_files = set(repaired.get("source_files", []))
    source_files.update(str(record.get("source_file", "")) for record in merged.get("raw_label_records", []) if record.get("source_file"))
    merged["source_files"] = sorted(source_files)
    merged["duke_core_labels"] = compact_core_labels(merged)
    return json_safe(merged)


def build_treatment_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "received_neoadjuvant_therapy": decode_received_neoadjuvant(row.get("Received Neoadjuvant Therapy or Not")),
        "neoadjuvant_chemotherapy": decode_yes_no(row.get("Neoadjuvant Chemotherapy")),
        "adjuvant_chemotherapy": decode_yes_no(row.get("Adjuvant Chemotherapy")),
        "neoadjuvant_endocrine_therapy": decode_yes_no(row.get("Neoadjuvant Endocrine Therapy Medications ")),
        "adjuvant_endocrine_therapy": decode_yes_no(row.get("Adjuvant Endocrine Therapy Medications ")),
        "neoadjuvant_anti_her2": decode_yes_no(row.get("Neoadjuvant Anti-Her2 Neu Therapy")),
        "adjuvant_anti_her2": decode_yes_no(row.get("Adjuvant Anti-Her2 Neu Therapy ")),
        "neoadjuvant_radiation": decode_yes_no(row.get("Neoadjuvant Radiation Therapy")),
        "adjuvant_radiation": decode_yes_no(row.get("Adjuvant Radiation Therapy")),
        "surgery": decode_yes_no(row.get("Surgery")),
        "surgery_type": decode_map(row.get("Definitive Surgery Type"), SURGERY_TYPE_MAP),
        "days_to_surgery_from_diagnosis": days_or_none(row.get("Days to Surgery (from the date of diagnosis)")),
    }


def build_survival_payload(row: dict[str, Any]) -> dict[str, Any]:
    local_recurrence_days = days_or_none(row.get("Days to local recurrence (from the date of diagnosis) "))
    distant_recurrence_days = days_or_none(row.get("Days to distant recurrence(from the date of diagnosis) "))
    death_days = days_or_none(row.get("Days to death (from the date of diagnosis) "))
    last_local_free_days = days_or_none(row.get("Days to last local recurrence free assessment (from the date of diagnosis) "))
    last_distant_free_days = days_or_none(row.get("Days to last distant recurrence free assemssment(from the date of diagnosis) "))
    last_contact_days = days_or_none(
        row.get(
            "Age at last contact in EMR f/u(days)(from the date of diagnosis) ,last time patient known to be alive, unless age of death is reported(in such case the age of death"
        )
    )
    event_times = [value for value in (local_recurrence_days, distant_recurrence_days, death_days) if value is not None]
    censor_times = [value for value in (last_local_free_days, last_distant_free_days, last_contact_days) if value is not None]
    dfs_event = 1 if event_times else 0 if censor_times else None
    dfs_time = min(event_times) if event_times else min(censor_times) if censor_times else None
    os_event = 1 if death_days is not None else 0 if last_contact_days is not None else None
    os_time = death_days if death_days is not None else last_contact_days
    recurrence_event = decode_yes_no(row.get("Recurrence event(s)"))
    return {
        "recurrence_event": recurrence_event,
        "local_recurrence_event": 1 if local_recurrence_days is not None else 0 if recurrence_event == "no" else None,
        "distant_recurrence_event": 1 if distant_recurrence_days is not None else 0 if recurrence_event == "no" else None,
        "death_event": os_event,
        "dfs_event": dfs_event,
        "dfs_time_days": dfs_time,
        "os_event": os_event,
        "os_time_days": os_time,
        "local_recurrence_days": local_recurrence_days,
        "distant_recurrence_days": distant_recurrence_days,
        "death_days": death_days,
        "last_local_recurrence_free_days": last_local_free_days,
        "last_distant_recurrence_free_days": last_distant_free_days,
        "last_contact_days": last_contact_days,
    }


def normalize_duke_pcr(value: Any) -> int | None:
    """Duke-specific pCR rule requested by the user.

    Pathologic Response to Neoadjuvant Therapy:
    1 -> pCR, any other non-missing code -> non-pCR, NA/blank -> missing.
    """

    if is_missing(value):
        return None
    number = int_or_none(value)
    if number is None:
        return None
    return 1 if number == 1 else 0


def decode_receptor(value: Any) -> str:
    return decode_map(value, RECEPTOR_MAP)


def decode_yes_no(value: Any) -> str:
    return decode_map(value, YES_NO_MAP)


def decode_received_neoadjuvant(value: Any) -> str:
    return decode_map(value, {1: "yes", 2: "no"})


def decode_map(value: Any, mapping: dict[int, str]) -> str:
    if is_missing(value):
        return "unknown"
    number = int_or_none(value)
    if number is None:
        return str(value).strip()
    return mapping.get(number, str(number))


def stage_summary(t_value: Any, n_value: Any, m_value: Any) -> dict[str, str]:
    t_stage = decode_stage(t_value, T_STAGE_MAP)
    n_stage = decode_stage(n_value, N_STAGE_MAP)
    m_stage = decode_stage(m_value, M_STAGE_MAP)
    parts = [part for part in (t_stage, n_stage, m_stage) if part != "unknown"]
    return {
        "t": t_stage,
        "n": n_stage,
        "m": m_stage,
        "summary": " ".join(parts) if parts else "unknown",
    }


def decode_stage(value: Any, mapping: dict[int, str]) -> str:
    if is_missing(value):
        return "unknown"
    number = int_or_none(value)
    if number is None:
        return str(value).strip()
    return mapping.get(number, str(number))


def normalize_laterality(value: Any) -> str:
    if is_missing(value):
        return "unknown"
    text = str(value).strip().upper()
    if text == "L" or text.startswith("L "):
        return "left"
    if text == "R" or text.startswith("R "):
        return "right"
    if text in {"B", "BILATERAL"}:
        return "bilateral"
    return str(value).strip()


def age_years_from_birth_days(value: Any) -> float | None:
    days = numeric_or_none(value)
    if days is None:
        return None
    return round(abs(float(days)) / 365.25, 2)


def days_or_none(value: Any) -> float | int | None:
    number = numeric_or_none(value)
    if number is None:
        return None
    return int(number) if float(number).is_integer() else number


def numeric_or_none(value: Any) -> float | int | None:
    if is_missing(value):
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else number


def int_or_none(value: Any) -> int | None:
    number = numeric_or_none(value)
    if number is None:
        return None
    return int(number)


def first_non_missing(*values: Any) -> Any:
    for value in values:
        if not is_missing(value):
            return clean_value(value)
    return None


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    if text == "":
        return True
    return text.upper() in MISSING_TOKENS


def clean_value(value: Any) -> Any:
    if is_missing(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_duke_subject_id(value: Any) -> str:
    if is_missing(value):
        return ""
    text = str(value).strip().replace("_", "-")
    if text.lower().startswith("breast-mri-"):
        prefix, number = text.rsplit("-", 1)
        if number.isdigit():
            return f"Breast-MRI-{int(number):03d}"
        return f"Breast-MRI-{number}"
    digits = "".join(char for char in text if char.isdigit())
    return f"Breast-MRI-{int(digits):03d}" if digits else text


def _unique_headers(values: Sequence[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values):
        base = str(value).strip() if value not in (None, "") else f"unnamed_{index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}.{count}")
    return headers


def _row_dict(headers: Sequence[str], values: Sequence[Any]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for key, value in zip(headers, values):
        if key.startswith("unnamed_"):
            continue
        output[key] = clean_value(value)
    return output


def is_duke_label_path(path: Path, dataset_id: str) -> bool:
    parts = {part.lower() for part in path.parts}
    if dataset_id.lower() in parts:
        return True
    existing = read_json_object(path)
    return existing.get("dataset_id") == dataset_id


def infer_subject_from_label_path(path: Path, dataset_id: str) -> str:
    parts = list(path.parts)
    lowered = [part.lower() for part in parts]
    if dataset_id.lower() in lowered:
        index = lowered.index(dataset_id.lower())
        if index + 1 < len(parts):
            return normalize_duke_subject_id(parts[index + 1])
    return ""


def label_summary_row(labels_path: Path, labels: dict[str, Any]) -> dict[str, Any]:
    return {
        "dataset_id": labels.get("dataset_id", DATASET_ID),
        "subject_id": labels.get("subject_id", ""),
        "timepoint": labels.get("timepoint", "unknown"),
        "study_uid": labels.get("study_uid", "unknown"),
        "labels_json": json.dumps(public_label_fields(labels), ensure_ascii=False, sort_keys=True),
        "pCR": labels.get("pCR"),
        "pcr_status": labels.get("pcr_status", "unknown"),
        "HER2": labels.get("HER2", "unknown"),
        "ER": labels.get("ER", "unknown"),
        "PR": labels.get("PR", "unknown"),
        "molecular_subtype": labels.get("molecular_subtype", "unknown"),
        "laterality": labels.get("laterality", "unknown"),
        "age": labels.get("age"),
        "BMI": labels.get("BMI"),
        "source_label_count": labels.get("source_label_count", 1),
        "split": labels.get("split", split_for_subject(labels.get("dataset_id", DATASET_ID), labels.get("subject_id", ""))),
    }


def update_row_with_public_labels(row: dict[str, str], labels: dict[str, Any]) -> None:
    public = public_label_fields(labels)
    row["labels_json"] = json.dumps(public, ensure_ascii=False, sort_keys=True)
    for key in ("pCR", "pcr_status", "HER2", "ER", "PR"):
        if key in row:
            row[key] = csv_value(labels.get(key))
    if "modalities" in row:
        row.setdefault("dataset_id", DATASET_ID)


def public_label_fields(labels: dict[str, Any]) -> dict[str, Any]:
    keys = [
        "pCR",
        "pcr_status",
        "HER2",
        "ER",
        "PR",
        "molecular_subtype",
        "age",
        "BMI",
        "sex",
        "laterality",
        "clinical_stage",
        "pathologic_stage",
        "tumor_size_cm",
        "treatment",
        "response",
        "survival",
    ]
    return {key: labels.get(key) for key in keys}


def compact_core_labels(labels: dict[str, Any]) -> dict[str, Any]:
    keys = [
        "pCR",
        "pcr_status",
        "HER2",
        "ER",
        "PR",
        "molecular_subtype",
        "age",
        "sex",
        "menopause",
        "race_ethnicity",
        "laterality",
        "clinical_stage",
        "pathologic_stage",
        "tumor_size_cm",
        "oncotype_score",
        "pathology",
        "treatment",
        "response",
        "survival",
    ]
    return {key: labels.get(key) for key in keys if known_value(labels.get(key))}


def known_value(value: Any) -> bool:
    if value in (None, "", "unknown"):
        return False
    if isinstance(value, dict):
        return any(known_value(item) for item in value.values())
    if isinstance(value, list):
        return any(known_value(item) for item in value)
    return True


def split_for_subject(dataset_id: str, subject_id: str, seed: int = 2026) -> str:
    import hashlib

    digest = hashlib.sha256(f"{seed}|{dataset_id}|{subject_id}".encode("utf-8")).hexdigest()
    value = int(digest[:8], 16) / 0xFFFFFFFF
    if value < 0.70:
        return "train"
    if value < 0.85:
        return "val"
    return "test"


def read_json_object(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return value if isinstance(value, dict) else {}


def write_json_atomic(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    tmp.replace(path)


def read_csv_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    if not path.exists():
        return [], []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader], list(reader.fieldnames or [])


def write_csv_atomic(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: csv_value(row.get(key, "")) for key in fieldnames})
    tmp.replace(path)


def csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [json_safe(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


if __name__ == "__main__":
    raise SystemExit(main())
